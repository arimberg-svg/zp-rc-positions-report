#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Сборка обезличенного отчёта ЗП РЦ по должностям.

Ожидает рядом (или через переменные окружения):
  ZP_DIR   — папка с ведомостями начисления (xlsx)
  TABEL_REPORT — путь к tabeli-rc/report.json с полем people[{tab, position}]
"""

from __future__ import annotations

import json
import os
import re
import statistics
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
ZP_DIR = Path(os.environ.get("ZP_DIR", ROOT.parent))
TABEL_REPORT = Path(
    os.environ.get(
        "TABEL_REPORT",
        r"C:\Users\User\Desktop\КУРСОРИО\Табели РЦ и Логистики\tabeli-rc\report.json",
    )
)
EXCLUDE = {"Руководитель распределительного центра и логистики"}
MONTHS_RU = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}
MONTH_LABEL = {
    "2025-08": "Авг 2025",
    "2025-09": "Сен 2025",
    "2025-10": "Окт 2025",
    "2025-11": "Ноя 2025",
    "2025-12": "Дек 2025",
    "2026-01": "Янв 2026",
    "2026-02": "Фев 2026",
    "2026-03": "Мар 2026",
    "2026-04": "Апр 2026",
    "2026-05": "Май 2026",
    "2026-06": "Июн 2026",
    "2026-07": "Июл 2026",
}


def parse_month(value: object) -> str | None:
    if not value:
        return None
    text = str(value).strip().lower()
    m = re.search(
        r"(январь|февраль|март|апрель|май|июнь|июль|август|сентябрь|октябрь|ноябрь|декабрь)\s+(\d{4})",
        text,
    )
    if not m:
        return None
    return f"{m.group(2)}-{MONTHS_RU[m.group(1)]:02d}"


def main() -> None:
    people = json.loads(TABEL_REPORT.read_text(encoding="utf-8"))["people"]
    tab_pos = {str(p["tab"]).strip(): p["position"] for p in people}

    zp: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for path in sorted(ZP_DIR.glob("*.xlsx")):
        if path.name.startswith("~") or path.parent.name == "github-report":
            continue
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        ym = parse_month(ws.cell(4, 6).value)
        if not ym:
            continue
        current_tab: str | None = None
        for row in range(8, ws.max_row + 1):
            row_vals = [ws.cell(row, c).value for c in range(1, 19)]
            if any(isinstance(v, str) and "итого" in v.lower() for v in row_vals if v):
                break
            tab = ws.cell(row, 9).value
            accrual = ws.cell(row, 11).value
            amount = ws.cell(row, 18).value
            if tab:
                current_tab = str(tab).strip()
            if current_tab and accrual and isinstance(amount, (int, float)):
                zp[ym][current_tab] += float(amount)

    by_pos_month: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
    pos_all: dict[str, list[float]] = defaultdict(list)
    monthly = []
    yms = sorted(zp)

    for ym in yms:
        vals: list[float] = []
        for tab, salary in zp[ym].items():
            pos = tab_pos.get(tab)
            if not pos or pos in EXCLUDE:
                continue
            vals.append(salary)
            by_pos_month[pos][ym].append(salary)
            pos_all[pos].append(salary)
        monthly.append(
            {
                "ym": ym,
                "label": MONTH_LABEL[ym],
                "people": len(vals),
                "avg": round(statistics.mean(vals)) if vals else 0,
                "median": round(statistics.median(vals)) if vals else 0,
                "total": round(sum(vals)) if vals else 0,
                "min": round(min(vals)) if vals else 0,
                "max": round(max(vals)) if vals else 0,
            }
        )

    positions = []
    for pos, vals in sorted(pos_all.items(), key=lambda item: -statistics.mean(item[1])):
        peaks = sorted(
            ((ym, statistics.mean(v), len(v)) for ym, v in by_pos_month[pos].items()),
            key=lambda item: -item[1],
        )
        series = []
        for ym in yms:
            chunk = by_pos_month[pos].get(ym, [])
            series.append(round(statistics.mean(chunk)) if chunk else None)
        positions.append(
            {
                "position": pos,
                "person_months": len(vals),
                "months_present": sum(1 for x in series if x is not None),
                "avg": round(statistics.mean(vals)),
                "median": round(statistics.median(vals)),
                "min": round(min(vals)),
                "max": round(max(vals)),
                "peak_ym": peaks[0][0],
                "peak_label": MONTH_LABEL[peaks[0][0]],
                "peak_avg": round(peaks[0][1]),
                "peak_n": peaks[0][2],
                "low_ym": peaks[-1][0],
                "low_label": MONTH_LABEL[peaks[-1][0]],
                "low_avg": round(peaks[-1][1]),
                "monthly_avg": series,
            }
        )

    all_vals = [s for vals in pos_all.values() for s in vals]
    peak_month = max(monthly, key=lambda item: item["avg"])
    low_month = min(monthly, key=lambda item: item["avg"])

    report = {
        "title": "ЗП РЦ по должностям",
        "org": "У Михалыча",
        "department": "Отдел складской и транспортной логистики",
        "period": "Август 2025 — Июль 2026",
        "notes": [
            "ФИО и табельные номера исключены.",
            "Должность «Руководитель распределительного центра и логистики» исключена из отчёта.",
            "Должности сопоставлены с табелями Т-13 по табельному номеру (в итоговом файле номера не публикуются).",
            "Суммы — начисления за месяц (оклад по часам + доплаты); строки «Итого» ведомостей не учитываются.",
        ],
        "summary": {
            "person_months": sum(p["person_months"] for p in positions),
            "positions": len(positions),
            "overall_avg": round(statistics.mean(all_vals)),
            "overall_median": round(statistics.median(all_vals)),
            "peak_month": peak_month,
            "low_month": low_month,
        },
        "months": [MONTH_LABEL[ym] for ym in yms],
        "yms": yms,
        "monthly": monthly,
        "positions": positions,
    }

    out = ROOT / "report_data.json"
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {out}")
    print(f"Peak month: {peak_month['label']} ({peak_month['avg']})")


if __name__ == "__main__":
    main()
